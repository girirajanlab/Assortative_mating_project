from openpyxl import load_workbook
import pandas as pd

# Read in data
# Load workbook and get list of available sheetnames
wb = load_workbook('Analysis_files/GeneDx_deidentified_kinship.xlsx', data_only=True)
ws_names = wb.sheetnames
# Remove the first sheetname because it is not useful for this analysis
ws_names.pop(0)

# Read in data for all CNVs and add to DataFrame
all_cnvs = pd.DataFrame()
for cnv in ws_names:
    ws = wb[cnv]
    df = pd.DataFrame(ws.values)
    df.drop(df.index[[0, 1, 2, 3, 4]], inplace=True)
    header = pd.Series([i.rstrip() for i in df.iloc[0]])
    df.rename(columns=header, inplace=True)
    df.drop(df.index[0], inplace=True)
    df['CNV'] = [cnv]*len(df.index)
    all_cnvs = pd.concat([all_cnvs, df])
print('Starting out:', all_cnvs.shape[0])

# Some families are duplicates - remove them
# Duplicates are from information provided by our collaborators
duplicate_fams = ['1q21.1dupFAM05', '1q21.1delFAM10', '15q11.2delFAM05', '15q11.2delFAM17', '15q11.2delFAM20',
                  '15q11.2delFAM24', '16p13.11dupFAM15', '16p13.11dupFAM18', '16p11.2SH2B1delFAM17', '16p11.2delFAM46']
all_cnvs = all_cnvs[~all_cnvs['ID'].isin(duplicate_fams)]
print('Removing duplicates:', all_cnvs.shape[0])

# Remove families with VERY low kinship values
# Very low values may indicate that 2 individuals are from different populations
# Based on the igures from the original KING paper (https://www.chen.kingrelatedness.com/publications/pdf/BI26_2867.pdf),
# I will use a threshold of -0.1
all_cnvs = all_cnvs[all_cnvs['Parental Kinship Values'] > -0.1]
print('Removing highly negative kinship values:', all_cnvs.shape[0])

# Extract Inheritence from Annotation column and add as new column
def inherit(row):
    annotation=row.Annotation
    if 'dn' in annotation:
        return 'DN'
    if 'mat' in annotation:
        return 'M'
    elif 'pat' in annotation:
        return 'P'
    else:
        return 'unknown'
all_cnvs['Inheritance'] = all_cnvs.apply(inherit, axis=1)
# Combine maternal and paternal inheritance into one group
inherited = {'P':'inherited', 'M':'inherited', 'DN':'de novo', 'unknown':'unknown'}
all_cnvs['Inherited'] = all_cnvs['Inheritance'].map(inherited)

# Annotate CNVs as clinically-variable or syndromic
# Definitions from Girirajan NEJM 2012
clinically_variable = ['1q21.1 dup', '1q21.1 del', '3q29 deletion', '15q11.2 deletion', '15q13.3 deletion',
                       '16p13.11 deletion', '16p13.11 duplication', '16p12.1 deletion', '16p11.2 SH2B1 deletion',
                      '16p11.2 SH2B1 duplication', '16p11.2 deletion', '16p11.2 duplication', '17q12 deletion',
                      '22q11.2 DiGeorgeVCFS deletion']

syndromic = ['Sotos deletion', '7q11.23 Williams del', '7q11.23 Williams dup', 'Prader_Willi Angelman deletion',
           'Smith-Magenis deletion', 'Smith-Magenis duplication', '17q21.31 deletion']

all_cnvs['Expression']=''
all_cnvs.loc[all_cnvs.CNV.isin(clinically_variable), 'Expression']='variably expressive'
all_cnvs.loc[all_cnvs.CNV.isin(syndromic), 'Expression']='syndromic'

# Separate CNVs by del/dup and region
all_cnvs['Del/Dup']='del'
all_cnvs.loc[all_cnvs.CNV.str.contains('dup'), 'Del/Dup']='dup'

# Save as file
all_cnvs.to_csv('Analysis_files/1_GeneDx_formatted_data.csv', index = False)